from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.validators import validate_email
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.contrib.auth.hashers import make_password, check_password
from .models import Admin, User, News, Contacts, RecJobs, Feedback, RecCources, Jobs, Courses
from django.http import JsonResponse
import json
import random
import requests
import openpyxl
import numpy as np
from sentence_transformers import SentenceTransformer,util
from sklearn.metrics.pairwise import cosine_similarity
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
import re


def admin_users(request):
    # Get current date and start of week (Monday)
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    
    # Calculate statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    new_users_this_week = User.objects.filter(
        date_joined__gte=start_of_week
    ).count()
    
    users = User.objects.all()
    
    context = {
        'users': users,
        'total_users': total_users,
        'active_users': active_users,
        'new_users_this_week': new_users_this_week,
    }
    
    return render(request, 'admin-users.html', context)

# Create your views here.
def index(request):
    return render(request, 'index.html')

def about(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'about.html', context)

def contact(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'contact.html', context)

def news(request):
    news = News.objects.all() 
    return render(request, 'news.html', {'news': news})

def feedback(request):
    feedbacks = Feedback.objects.all() 
    return render(request, 'feedback.html', {'feedbacks': feedbacks})

def faq(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'faq.html', context)

def adminlogin(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'admin-login.html', context)

def userlogin(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'user-login.html', context)

def userforgot(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'user-forgot.html', context)

def userregister(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'user-register.html', context)


# FIX: admin_login (Handles MultipleObjectsReturned)
def admin_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        
        # Use filter().first() to safely retrieve the admin
        admin = Admin.objects.filter(username=username, password=password).first()
        
        if admin:
            request.session['admin_id'] = admin.id
            return redirect('/admin-home/')
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'admin-login.html')

def admin_logout(request):
    request.session.flush()
    return redirect('/admin-login/')

def adminhome(request):
    if 'admin_id' not in request.session:
        return redirect('/admin-login/')

    return render(request, 'admin-home.html')

# FIX: adminprofile (Handles MultipleObjectsReturned)
def adminprofile(request):
    if 'admin_id' not in request.session:
        return redirect('/admin-login/')
    
    adminid = request.session.get('admin_id')
    
    # Use filter().first() to safely retrieve the admin
    admin = Admin.objects.filter(id=adminid).first()
    
    if not admin:
        messages.error(request, "Admin user not found.")
        return redirect('/admin-login/') 

    return render(request, 'admin-profile.html',  {'admin': admin})

def adminnews(request):
    news = News.objects.all() 
    return render(request, 'admin-news.html',  {'news': news})

def adminusers(request):
    users = User.objects.all() 
    return render(request, 'admin-users.html',  {'users': users})

def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return redirect('admin-users')


def store_user(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        password = request.POST.get("password")
        otp = request.POST.get("otp")

        session_otp = request.session.get('otp')
        if str(otp) != str(session_otp):
            messages.error(request, "Invalid OTP!")
            return redirect("user-register")
        
        # Validations
        if not name or len(name) < 3:
            messages.error(request, "Name must be at least 3 characters long.")
            return redirect("user-register")

        if not phone.isdigit() or len(phone) < 10:
            messages.error(request, "Phone number must be at least 10 digits long.")
            return redirect("user-register")

        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Invalid email format.")
            return redirect("user-register")

        if len(password) < 6:
            messages.error(request, "Password must be at least 6 characters long.")
            return redirect("user-register")

        # Check if user already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect("user-register")

        if User.objects.filter(phone=phone).exists():
            messages.error(request, "Phone number already registered.")
            return redirect("user-register")

        # Hash password before storing - This is the correct, secure method
        hashed_password = make_password(password)

        # Save user to database
        user = User(name=name, phone=phone, email=email, password=hashed_password)
        user.save()

        if 'otp' in request.session:
             del request.session['otp']

        messages.success(request, "User registered successfully!")
        return redirect("user-register")

    return render(request, "user-register.html")

def update_user(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        
        # Validations
        if not name or len(name) < 3:
            messages.error(request, "Name must be at least 3 characters long.")
            return redirect("user-profile")

        if not phone.isdigit() or len(phone) < 10:
            messages.error(request, "Phone number must be at least 10 digits long.")
            return redirect("user-profile")

        # Save user to database
        user_id = request.session.get('user_id')
        user = get_object_or_404(User, id=user_id)
        user.name = name
        user.phone = phone
        user.save()

        messages.success(request, "User details updated successfully!")
        return redirect("user-profile")

    return render(request, "user-profile.html")

def user_forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Invalid email format.")
            return redirect("user-forgot")
        
        user = User.objects.filter(email=email).first()
        if user:
            # Generate OTP and store in session
            otp = random.randint(100000, 999999)
            request.session['reset_otp'] = otp
            request.session['reset_user_id'] = user.id
            request.session.set_expiry(300)  # 5 minutes expiry

            # Send email with OTP
            url = f"https://threeartisans.com/sendmail.php?type=jobs&email={email}&full_name={user.name}&otp={otp}&subject=Password+Reset+Request"
            
            try:
                response = requests.get(url, verify=False)
                if response.status_code == 200:
                    messages.success(request, f"OTP has been sent to your email: {email}")
                    # REDIRECT TO RESET PAGE WITH USER ID
                    return redirect('user-reset', user_id=user.id)
                else:
                    messages.error(request, "Failed to send OTP. Please try again.")
            except Exception as e:
                messages.error(request, "Email service temporarily unavailable. Please try again later.")
        else:
            messages.error(request, "Email address not found in our system.")
        
        return redirect("user-forgot")
    
    return render(request, 'user-forgot.html')
def user_reset(request, user_id):
    # Make sure user_id is passed to template
    return render(request, 'user-reset.html', {'user_id': user_id})

def user_reset_password(request, user_id=None):
    if request.method == "GET":
        # Show the reset password form
        return render(request, 'user-reset.html', {'user_id': user_id})
    
    elif request.method == "POST":
        otp = request.POST.get("otp")
        newpass = request.POST.get("newpass")
        confpass = request.POST.get("confpass")
        user_id = request.POST.get("userid") or user_id

        if not user_id:
            messages.error(request, "Invalid reset request.")
            return redirect("user-forgot")

        # Verify OTP
        session_otp = request.session.get('reset_otp')
        session_user_id = request.session.get('reset_user_id')
        
        if not session_otp or str(otp) != str(session_otp) or int(user_id) != session_user_id:
            messages.error(request, "Invalid or expired OTP!")
            return render(request, 'user-reset.html', {'user_id': user_id})
      
        # Validate password
        if len(newpass) < 6:
            messages.error(request, "Password must be at least 6 characters long.")
            return render(request, 'user-reset.html', {'user_id': user_id})
        
        if newpass != confpass:
            messages.error(request, "Passwords do not match.")
            return render(request, 'user-reset.html', {'user_id': user_id})

        try:
            # Update password
            user = User.objects.get(id=user_id)
            hashed_password = make_password(newpass)
            user.password = hashed_password
            user.save()

            # Clear reset session data
            if 'reset_otp' in request.session:
                del request.session['reset_otp']
            if 'reset_user_id' in request.session:
                del request.session['reset_user_id']

            messages.success(request, "Password reset successfully! You can now login with your new password.")
            return redirect("user-login")
            
        except User.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect("user-forgot")
    
    return redirect("user-forgot")

def user_password_change(request):
    if request.method == "POST":
        currpass = request.POST.get("currpass")
        newpass = request.POST.get("newpass")
        confpass = request.POST.get("confpass")

        user_id = request.session.get('user_id')
        user = get_object_or_404(User, id=user_id)
        
        if check_password(currpass, user.password):
            pass
        else:
            messages.error(request, "Current password not exist.")
            return redirect("user-profile")
        
        if len(newpass) < 6:
            messages.error(request, "Password must be at least 6 characters long.")
            return redirect("user-profile")
        
        if newpass != confpass:
            messages.error(request, "Passwords not matched.")
            return redirect("user-profile")

        hashed_password = make_password(newpass)
        user.password = hashed_password
        user.save()

        messages.success(request, "User password changed successfully!")
        return redirect("user-profile")

    return render(request, "user-profile.html")

def store_news(request):
    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")

        # Validations
        if not title or len(title) < 3:
            messages.error(request, "Title must be at least 3 characters long.")
            return redirect("admin-news")

        news = News(title=title, description=description)
        news.save()

        messages.success(request, "News added successfully!")
        return redirect("admin-news")

    return render(request, "admin-news.html")

def delete_news(request, news_id):
    news = get_object_or_404(News, id=news_id)
    news.delete()
    return redirect('admin-news')

def sendHtmlMail(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            content= data.get("content", "")
            user_id = request.session.get('user_id')
            user = get_object_or_404(User, id=user_id)
            email = user.email

            payload = {"type": "jobs", "email": email,"full_name":email,"subject" : "Recommended Jobs","content":content}
            headers ={"Accept":"application/json"}
            
            # Note: Added verify=False for consistency/SSL bypass
            response = requests.post("https://threeartisans.com/sendmail.php", data=payload, verify=False)
            
            if response.status_code == 200:
                return JsonResponse({"message": "Mail send", "status": "success"}, status=200)
            else:
                 return JsonResponse({"message": "External mail service failed.", "status": "error"}, status=500)
        
        except Exception as e:
            print(f"Error: {str(e)}") 
            return JsonResponse({"message": "An unexpected server error occurred.", "status":str(e)}, status=500)

# FIX: sendMail (Now returns OTP in JSON response)
def sendMail(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            email = data.get("email", "")

            # 1. Validate email
            validate_email(email)
            
            # 2. Check if user already exists
            if User.objects.filter(email=email).exists():
                 return JsonResponse({"message": "Email already registered.", "status": "error"}, status=409)

            otp = random.randint(100000, 999999)

            request.session['otp'] = otp
            request.session.set_expiry(300)

            # 3. External API call for sending email
            url= "https://threeartisans.com/sendmail.php?type=jobs&email="+email+"&full_name="+email+"&otp="+str(otp)+"&subject=New+OTP+for+verification"
            
            # FIX: Add verify=False to bypass SSL certificate error
            response = requests.get(url, verify=False)

            if response.status_code == 200:
                # MODIFICATION: Return the generated OTP in the response for on-screen display/debugging
                return JsonResponse({"message": "OTP sent at your email id.", "status": "success", "otp_code": otp}, status=200)
            else:
                 return JsonResponse({"message": "Failed to send OTP via external service.", "status": "error"}, status=500)

        except ValidationError:
            return JsonResponse({"message": "Invalid email format.", "status": "error"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"message": "Invalid data format (JSON expected).", "status": "error"}, status=400)
        
        except Exception as e:
            print(f"Error in sendMail: {e}")
            return JsonResponse({"message": "An unexpected server error occurred.", "status": "error"}, status=500)


    return JsonResponse({"message": "Invalid request method", "status": "error"}, status=405)

# FIX: user_login (Handles DoesNotExist and the plain-text password issue)
def user_login(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        if not email or not password:
             messages.error(request, "Please enter both email and password.")
             return render(request, 'user-login.html')

        try:
            user = User.objects.get(email=email) 
            
            # SECURE/CORRECT METHOD (Checks against a hashed password)
            is_valid_password = check_password(password, user.password)

            # TEMPORARY FALLBACK FOR PLAIN-TEXT DB ENTRIES (Insecure, remove after fixing DB)
            if not is_valid_password and len(user.password) < 10: 
                if password == user.password:
                    is_valid_password = True
                    # Highly recommended: user.password = make_password(password); user.save()

            if is_valid_password:
                request.session['user_id'] = user.id
                request.session['user_name'] = user.name
                messages.success(request, f"Welcome back, {user.name}!")
                return redirect('/user-home/')
            else:
                messages.error(request, "Invalid email or password.")
                
        # FIX: Catch the DoesNotExist exception
        except User.DoesNotExist:
            messages.error(request, "Invalid email or password.")

        return render(request, 'user-login.html') 

    return render(request, 'user-login.html') 


def userhome(request):
    if 'user_id' not in request.session:
        return redirect('/user-login/')
    news = News.objects.all() 
    return render(request, 'user-home.html', {'news': news})
    
def userprofile(request):
    if 'user_id' not in request.session:
        return redirect('/user-login/')
    
    userid = request.session.get('user_id')
    user = get_object_or_404(User, id=userid)
    return render(request, 'user-profile.html',  {'user': user})

def usernews(request):
    news = News.objects.all() 
    return render(request, 'user-news.html',  {'news': news})

def userfeedback(request):
    feedbacks = Feedback.objects.select_related('user').all() 
    return render(request, 'user-feedback.html',  {'feedbacks': feedbacks})

def userfaq(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'user-faq.html', context)

def user_logout(request):
    request.session.flush()
    return redirect('/user-login/')

def findjob(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'find-job.html', context)

def findcourse(request):
    context = {'title': 'Django Home', 'message': 'Hello from Django!'}
    return render(request, 'find-course.html', context)

# Initialize embedding model once
model = SentenceTransformer('all-MiniLM-L6-v2')


def clean_text(text):
    """Normalize and clean text for better matching."""
    text = re.sub(r'[^a-zA-Z0-9\s]', '', text)
    return text.lower().strip()


# ======================================================
# 🔹 JOB RECOMMENDATION FUNCTION
# ======================================================
def findjobs(request):
    if request.method == "POST":
        # Remove the AI section completely
        interest = clean_text(request.POST.get("interest", ""))
        job_category = clean_text(request.POST.get("job_category", ""))
        experience = clean_text(request.POST.get("experience", ""))
        skills = clean_text(request.POST.get("skills", ""))
        mode = clean_text(request.POST.get("mode", ""))

        try:
            all_jobs = RecJobs.objects.all()
            if not all_jobs.exists():
                messages.error(request, "No jobs found in database.")
                return redirect("find-job")

            job_texts, job_ids = [], []
            for job in all_jobs:
                job_texts.append(clean_text(f"{job.job_role}. {job.description}"))
                job_ids.append(job.id)

            # Filter roughly first
            filtered_jobs = [
                (job_ids[i], job_texts[i])
                for i in range(len(job_texts))
                if any(skill in job_texts[i] for skill in skills.split(","))
            ] or list(zip(job_ids, job_texts))

            job_embeddings = model.encode([j[1] for j in filtered_jobs])
            query_text = f"{interest} {job_category} {skills} {experience} {mode}"
            query_embedding = model.encode([query_text])

            similarity_scores = cosine_similarity(query_embedding, job_embeddings)[0]
            sorted_indices = similarity_scores.argsort()[::-1]

            results = []
            for idx in sorted_indices[:5]:
                score = float(similarity_scores[idx])
                job = RecJobs.objects.get(id=filtered_jobs[idx][0])
                results.append({
                    "job_title": job.job_role,
                    "description": job.description,
                    "similarity_score": round(score, 2),
                })

            return render(request, "show-db-jobs.html", {
                "jobs": results,
                "query": query_text
            })

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect("find-job")

    return render(request, "find-job.html")


# ======================================================
# 🔹 COURSE RECOMMENDATION FUNCTION
# ======================================================
def findcourses(request):
    if request.method == "POST":
        interest = clean_text(request.POST.get("interest", ""))
        course_type = clean_text(request.POST.get("course_type", ""))
        experience = clean_text(request.POST.get("experience_level", ""))
        budget = clean_text(request.POST.get("budget", ""))
        mode = clean_text(', '.join(request.POST.getlist("mode", [])))
        skills = clean_text(request.POST.get("skills", ""))

        try:
            all_courses = Courses.objects.all()
            if not all_courses.exists():
                messages.error(request, "No courses found in database.")
                return redirect("find-course")

            # Prepare course texts
            course_texts, course_ids = [], []
            for c in all_courses:
                text = clean_text(
                    f"{c.suggested_course} | {c.interest} | {c.course_type} | {c.experience_level} | {c.skills} | {c.mode} | {c.description}"
                )
                course_texts.append(text)
                course_ids.append(c.id)

            # Filter by skill keyword first
            filtered = [
                (course_ids[i], course_texts[i])
                for i in range(len(course_texts))
                if any(skill in course_texts[i] for skill in skills.split(","))
            ] or list(zip(course_ids, course_texts))

            embeddings = model.encode([t[1] for t in filtered], convert_to_tensor=True)
            user_query = f"{interest} {course_type} {experience} {budget} {mode} {skills}"
            query_embedding = model.encode(user_query, convert_to_tensor=True)

            sim_scores = util.cos_sim(query_embedding, embeddings)[0].cpu().numpy()
            top_indices = np.argsort(sim_scores)[::-1]

            results = []
            for i in top_indices[:5]:
                score = float(sim_scores[i])
                course = Courses.objects.get(id=filtered[i][0])
                results.append({
                    'suggested_course': course.suggested_course,
                    'description': course.description or f"Best suited for {course.experience_level} learners.",
                    'skills': course.skills or f"Focus on {interest}",
                    'topics': f"{course.course_type}, {course.mode}",
                    'provider': course.provider or "Internal",
                    'similarity_score': round(score, 2),
                })

            return render(request, 'show-db-courses.html', {'courses': results})

        except Exception as e:
            messages.error(request, f"Database search error: {e}")
            return redirect("find-course")

    return render(request, 'find-course.html')

def admincontacts(request):
    contacts = Contacts.objects.all() 
    return render(request, 'admin-contacts.html',  {'contacts': contacts})

def delete_contact(request, contact_id):
    con = get_object_or_404(Contacts, id=contact_id)
    con.delete()
    return redirect('admin-contacts')

def store_contact(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        message = request.POST.get("message")
        
        # Validations
        if not name or len(name) < 3:
            messages.error(request, "Name must be at least 3 characters long.")
            return redirect("contact")

        if not phone.isdigit() or len(phone) < 10:
            messages.error(request, "Phone number must be at least 10 digits long.")
            return redirect("contact")

        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Invalid email format.")
            return redirect("contact")

        if len(message) < 6:
            messages.error(request, "Message must be at least 6 characters long.")
            return redirect("contact")

        # Check if user already exists
        if Contacts.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect("contact")

        con = Contacts(name=name, phone=phone, email=email, message=message)
        con.save()

        messages.success(request, "Contact form submitted successfully!")
        return redirect("contact")

    return render(request, "contact.html")

def feedback_store(request):
    if request.method == "POST":
        description = request.POST.get("feedback")
        userid = request.session.get('user_id')
        # Validations
        if not description or len(description) < 10:
            messages.error(request, "Feedback must be at least 10 characters long.")
            return redirect("user-feedback")

        feedback = Feedback(user_id=userid, description=description)
        feedback.save()

        messages.success(request, "Feedback added successfully!")
        return redirect("user-feedback")

    return render(request, "user-feedback.html")

def admin_password_change(request):
    if request.method == "POST":
        currpass = request.POST.get("currpass")
        newpass = request.POST.get("newpass")
        confpass = request.POST.get("confpass")

        user_id = request.session.get('admin_id')
        user = get_object_or_404(Admin, id=user_id)
        
        if currpass == user.password:
            pass
        else:
            messages.error(request, "Current password not exist.")
            return redirect("admin-profile")
        
        if len(newpass) < 6:
            messages.error(request, "Password must be at least 6 characters long.")
            return redirect("admin-profile")
        
        if newpass != confpass:
            messages.error(request, "Passwords not matched.")
            return redirect("admin-profile")

        user.password = newpass
        user.save()

        messages.success(request, "User password changed successfully!")
        return redirect("admin-profile")

    return render(request, "admin-profile.html")

def admin_upload_jobs(request):
    if request.method == 'POST' and request.FILES.get('jobs'):
        excel_jobs = request.FILES['jobs']
        wb = openpyxl.load_workbook(excel_jobs)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Columns: job_role, description, category, skills, experience_level
            if len(row) < 5:
                continue  # Skip incomplete rows
            job_role, description, category, skills, experience_level = row

            if not RecJobs.objects.filter(
                job_role=job_role,
                description=description,
                category=category,
                skills=skills,
                experience_level=experience_level
            ).exists():
                RecJobs.objects.create(
                    job_role=job_role,
                    description=description,
                    category=category,
                    skills=skills,
                    experience_level=experience_level
                )

        messages.success(request, "Jobs uploaded successfully!")
        return redirect('admin-profile')
    
    return render(request, 'admin-profile.html')
def admin_upload_courses(request):
    if request.method == 'POST' and request.FILES.get('courses'):
        excel_courses = request.FILES['courses']
        wb = openpyxl.load_workbook(excel_courses)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Columns: interest, course_type, experience_level, mode, budget, duration,
            # suggested_course, description, skills, topics, provider, rating
            if len(row) < 12:
                continue  # Skip incomplete rows

            (
                interest, course_type, experience_level, mode, budget, duration,
                suggested_course, description, skills, topics, provider, rating
            ) = row

            if not Courses.objects.filter(
                interest=interest,
                course_type=course_type,
                experience_level=experience_level,
                suggested_course=suggested_course
            ).exists():
                Courses.objects.create(
                    interest=interest,
                    course_type=course_type,
                    experience_level=experience_level,
                    mode=mode,
                    budget=budget,
                    duration=duration,
                    suggested_course=suggested_course,
                    description=description,
                    skills=skills,
                    topics=topics,
                    provider=provider,
                    rating=rating or 0.0
                )

        messages.success(request, "Courses uploaded successfully!")
        return redirect('admin-profile')
    
    return render(request, 'admin-profile.html')

def adminhome(request):
    if 'admin_id' not in request.session:
        return redirect('/admin-login/')

    # Get counts from database
    total_users = User.objects.count()
    total_jobs = RecJobs.objects.count()
    total_courses = Courses.objects.count()
    total_contacts = Contacts.objects.count()
    total_feedback = Feedback.objects.count()
    total_news = News.objects.count()

    context = {
        'total_users': total_users,
        'total_jobs': total_jobs,
        'total_courses': total_courses,
        'total_contacts': total_contacts,
        'total_feedback': total_feedback,
        'total_news': total_news,
    }

    return render(request, 'admin-home.html', context)
